"""Manage students: list/search, create, delete, issue certificates, and
bulk Excel/CSV import (now supports an explicit Password column, falling
back to the default Gurukul@<MemberID> scheme when one isn't given)."""
import csv
import io
import re
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db import get_db
from app.deps import get_current_user, require_admin, CurrentUser
from app.models_orm import Student, User, AuditLog, Certificate, Notification
from app.schemas import StudentCreate, CertificateCreate
from app.security import hash_password
from app.serialize import serialize, serialize_many
from app.student_auth import default_student_password

router = APIRouter(prefix="/api/students", tags=["students"])

HEADER_MAP = {
    "member id": "member_id",
    "admission no": "admission_no",
    "admission number": "admission_no",
    "student name": "name",
    "name": "name",
    "password": "password",
    "class": "class",
    "section": "section",
    "house": "house",
    "room number": "room_number",
    "room": "room_number",
    "phone number": "phone",
    "phone": "phone",
    "gender": "gender",
    "hostel": "hostel",
    "photo url": "photo_url",
    "date of birth": "date_of_birth",
    "dob": "date_of_birth",
    "parent name": "parent_name",
    "parent phone": "parent_phone",
    "email": "email",
    "blood group": "blood_group",
    "address": "address",
    "joining date": "joining_date",
}

STUDENT_FIELDS = {
    "member_id", "admission_no", "name", "class", "section", "house", "room_number",
    "phone", "gender", "hostel", "photo_url", "date_of_birth", "parent_name",
    "parent_phone", "email", "blood_group", "address", "joining_date",
}


@router.get("")
def list_students(q: Optional[str] = None, user: CurrentUser = Depends(get_current_user), db: Session = Depends(get_db)):
    query = db.query(Student).order_by(Student.total_points.desc())
    students = query.all()
    if q:
        ql = q.strip().lower()
        if ql:
            def matches(s: Student) -> bool:
                fields = [s.member_id, s.name, s.admission_no, s.phone, s.room_number, s.house]
                return any(f and ql in str(f).lower() for f in fields)
            students = [s for s in students if matches(s)]
    return serialize_many(students)


@router.post("")
def add_student(payload: StudentCreate, user: CurrentUser = Depends(require_admin), db: Session = Depends(get_db)):
    member_id = payload.member_id.strip()
    name = payload.name.strip()
    if not member_id or not name:
        raise HTTPException(status_code=400, detail="Member ID and Name are required")
    if db.query(Student).filter(Student.member_id == member_id).first():
        raise HTTPException(status_code=400, detail=f"Member ID {member_id} already exists")

    data = payload.model_dump(by_alias=True, exclude={"password"}, exclude_none=True)
    if "class" in data:
        data["class_"] = data.pop("class")
    student = Student(**data)
    db.add(student)
    db.flush()

    # Always provision a login so the student can access their dashboard,
    # same fallback scheme used by bulk import when no password is given.
    effective_password = (payload.password or "").strip() or default_student_password(member_id)
    user_acc = User(hashed_password=hash_password(effective_password), full_name=name, role="student")
    db.add(user_acc)
    db.flush()
    student.user_id = user_acc.id

    db.commit()
    db.refresh(student)
    result = serialize(student)
    result["generated_password"] = effective_password
    return result


@router.delete("/{student_id}")
def delete_student(student_id: str, user: CurrentUser = Depends(require_admin), db: Session = Depends(get_db)):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    member_id, name, linked_user_id = student.member_id, student.name, student.user_id
    db.delete(student)
    if linked_user_id:
        linked_user = db.query(User).filter(User.id == linked_user_id).first()
        if linked_user:
            db.delete(linked_user)
    db.add(AuditLog(user_id=user.id, actor_name=user.full_name, action="delete_student", details=f"Deleted {member_id} ({name})"))
    db.commit()
    return {"ok": True}


@router.post("/{student_id}/certificate")
def issue_certificate(student_id: str, payload: CertificateCreate, user: CurrentUser = Depends(require_admin), db: Session = Depends(get_db)):
    if not payload.title.strip():
        raise HTTPException(status_code=400, detail="Certificate title is required")
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    cert = Certificate(student_id=student_id, title=payload.title.strip(), description=payload.description, issued_by=user.full_name)
    db.add(cert)
    db.add(Notification(student_id=student_id, title="Certificate Earned!", message=payload.title.strip(), type="certificate"))
    db.commit()
    db.refresh(cert)
    return serialize(cert)


def _provision_login(db: Session, student: Student, password: str, name: str) -> None:
    user_acc = User(hashed_password=hash_password(password), full_name=name, role="student")
    db.add(user_acc)
    db.flush()
    student.user_id = user_acc.id


def _parse_rows(filename: str, content: bytes) -> list[dict]:
    rows: list[dict] = []
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        raw_rows = list(csv.DictReader(io.StringIO(text)))
    else:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        raw_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_rows.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})

    for raw in raw_rows:
        out = {}
        for key, val in raw.items():
            if key is None:
                continue
            field = HEADER_MAP.get(str(key).strip().lower())
            if field:
                out[field] = "" if val is None else str(val).strip()
        if out.get("member_id") or out.get("name"):
            rows.append(out)
    return rows


@router.post("/import")
def import_students(file: UploadFile = File(...), user: CurrentUser = Depends(require_admin), db: Session = Depends(get_db)):
    content = file.file.read()
    try:
        rows = _parse_rows(file.filename or "", content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the file: {e}")

    imported = updated = skipped = 0
    failures: list[dict] = []

    for i, raw in enumerate(rows):
        row_num = i + 2
        r = {k: v for k, v in raw.items() if k in STUDENT_FIELDS and v}
        member_id = raw.get("member_id", "")
        name = raw.get("name", "")
        password = raw.get("password", "").strip()

        if not member_id or not name:
            if not member_id and not name:
                skipped += 1
                continue
            failures.append({"row": row_num, "member_id": member_id, "reason": "member_id and name are required"})
            continue

        phone = r.get("phone")
        if phone and not re.match(r"^\d+$", phone):
            failures.append({"row": row_num, "member_id": member_id, "reason": "Phone must be digits only"})
            continue

        admission_no = r.get("admission_no")
        if admission_no:
            conflict = (
                db.query(Student)
                .filter(Student.admission_no == admission_no, Student.member_id != member_id)
                .first()
            )
            if conflict:
                failures.append({
                    "row": row_num, "member_id": member_id,
                    "reason": f"Admission No {admission_no} already used by {conflict.member_id}",
                })
                continue

        existing = db.query(Student).filter(Student.member_id == member_id).first()
        effective_password = password or default_student_password(member_id)

        if existing:
            for k, v in r.items():
                setattr(existing, k if k != "class" else "class_", v)
            if not existing.user_id:
                _provision_login(db, existing, effective_password, name)
            elif password:
                # explicit password in the sheet updates an existing login too
                user_acc = db.query(User).filter(User.id == existing.user_id).first()
                if user_acc:
                    user_acc.hashed_password = hash_password(password)
            updated += 1
        else:
            data = {(k if k != "class" else "class_"): v for k, v in r.items()}
            student = Student(**data)
            db.add(student)
            db.flush()
            _provision_login(db, student, effective_password, name)
            imported += 1

    db.add(AuditLog(
        user_id=user.id, actor_name=user.full_name, action="excel_import",
        details=f"Imported {imported}, Updated {updated}, Skipped {skipped}, Failed {len(failures)}",
    ))
    db.commit()

    return {
        "total": len(rows),
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "failed": len(failures),
        "failures": failures,
    }
