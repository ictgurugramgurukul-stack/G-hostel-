"""Daily room-inspection checklist.

Every day a teacher ticks four boxes per student's room - Bed Arrangement,
Cupboard, Cleanliness, Blanket Folded - and it's saved as one row per
student per day (re-ticking the same day updates that row instead of
duplicating it). Admins can browse recent records and download the full
history as an Excel file.
"""
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.db import get_db
from app.deps import require_admin, require_staff, CurrentUser
from app.models_orm import AuditLog, PointTransaction, RoomInspection, Student
from app.schemas import InspectionSubmit

router = APIRouter(prefix="/api/inspections", tags=["inspections"])

# Every unticked checklist item costs the student this many points,
# deducted automatically the moment the box is left unchecked - no separate
# "Activity" needs to be awarded by hand for it.
AUTO_DEDUCT_PER_ITEM = 3
CHECKLIST_LABELS = {
    "bed_arrangement": "Bed Arrangement Not Done",
    "cupboard": "Cupboard Not Organized",
    "cleanliness": "Cleanliness Not Done",
    "blanket_folded": "Blanket Not Folded",
}
CHECKLIST_KEYS = ("bed_arrangement", "cupboard", "cleanliness", "blanket_folded")


def _today() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


@router.get("/today")
def today_checklist(date: Optional[str] = None, user: CurrentUser = Depends(require_staff), db: Session = Depends(get_db)):
    """Every student, plus today's (or the given date's) checklist state for
    each - unticked students simply come back with everything false."""
    the_date = (date or "").strip() or _today()

    students = db.query(Student).order_by(Student.room_number, Student.name).all()
    records = {
        r.student_id: r
        for r in db.query(RoomInspection).filter(RoomInspection.date == the_date).all()
    }

    rows = []
    for s in students:
        r = records.get(s.id)
        rows.append({
            "student_id": s.id,
            "name": s.name,
            "member_id": s.member_id,
            "room_number": s.room_number,
            "class": s.class_,
            "section": s.section,
            "house": s.house,
            "bed_arrangement": bool(r.bed_arrangement) if r else False,
            "cupboard": bool(r.cupboard) if r else False,
            "cleanliness": bool(r.cleanliness) if r else False,
            "blanket_folded": bool(r.blanket_folded) if r else False,
            "remarks": (r.remarks if r else "") or "",
            "recorded": r is not None,
        })

    return {"date": the_date, "students": rows}


@router.post("")
def submit_checklist(payload: InspectionSubmit, user: CurrentUser = Depends(require_staff), db: Session = Depends(get_db)):
    student = db.query(Student).filter(Student.id == payload.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    the_date = (payload.date or "").strip() or _today()

    record = (
        db.query(RoomInspection)
        .filter(RoomInspection.student_id == payload.student_id, RoomInspection.date == the_date)
        .first()
    )
    if not record:
        record = RoomInspection(student_id=payload.student_id, date=the_date)
        db.add(record)

    record.bed_arrangement = payload.bed_arrangement
    record.cupboard = payload.cupboard
    record.cleanliness = payload.cleanliness
    record.blanket_folded = payload.blanket_folded
    record.remarks = (payload.remarks or "").strip() or None
    record.teacher_id = user.id
    record.teacher_name = user.full_name

    # ---- auto-deduct for anything left unticked (idempotent) ----
    # Re-saving the same day's checklist must never deduct twice, so we
    # only ever charge the *difference* between what's already been
    # deducted for this record and what should be deducted now.
    unchecked = [k for k in CHECKLIST_KEYS if not getattr(record, k)]
    target_deduction = AUTO_DEDUCT_PER_ITEM * len(unchecked)
    already_deducted = record.auto_deducted_points or 0
    delta = already_deducted - target_deduction  # negative = deduct more, positive = refund

    if delta != 0:
        student.total_points += delta
        db.add(PointTransaction(
            student_id=student.id,
            activity_id=None,
            teacher_id=user.id,
            teacher_name=user.full_name,
            activity_name="Room Inspection: " + (", ".join(CHECKLIST_LABELS[k] for k in unchecked) if unchecked else "All Items Done"),
            points=delta,
            remarks=f"Auto-{'deducted' if delta < 0 else 'restored'} from room inspection on {the_date}",
        ))
        record.auto_deducted_points = target_deduction

    db.add(AuditLog(
        user_id=user.id, actor_name=user.full_name, action="room_inspection",
        details=f"{student.name} ({student.member_id}) \u2013 {the_date}",
    ))

    db.commit()
    db.refresh(record)

    return {
        "ok": True,
        "date": the_date,
        "bed_arrangement": record.bed_arrangement,
        "cupboard": record.cupboard,
        "cleanliness": record.cleanliness,
        "blanket_folded": record.blanket_folded,
        "auto_deducted_points": record.auto_deducted_points,
    }


def _query_records(db: Session, start_date: Optional[str], end_date: Optional[str], room: Optional[str]):
    query = db.query(RoomInspection)
    if start_date:
        query = query.filter(RoomInspection.date >= start_date)
    if end_date:
        query = query.filter(RoomInspection.date <= end_date)
    query = query.order_by(RoomInspection.date.desc())
    records = query.limit(2000).all()
    if room:
        room_l = room.strip().lower()
        student_ids = {s.id for s in db.query(Student).filter(Student.room_number.ilike(f"%{room_l}%")).all()}
        records = [r for r in records if r.student_id in student_ids]
    return records


@router.get("/history")
def history(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    room: Optional[str] = None,
    limit: int = 100,
    user: CurrentUser = Depends(require_admin),
    db: Session = Depends(get_db),
):
    records = _query_records(db, start_date, end_date, room)[:limit]
    students = {s.id: s for s in db.query(Student).all()}

    out = []
    for r in records:
        s = students.get(r.student_id)
        out.append({
            "date": r.date,
            "student_name": s.name if s else "(deleted student)",
            "member_id": s.member_id if s else "",
            "room_number": s.room_number if s else "",
            "bed_arrangement": r.bed_arrangement,
            "cupboard": r.cupboard,
            "cleanliness": r.cleanliness,
            "blanket_folded": r.blanket_folded,
            "remarks": r.remarks or "",
            "teacher_name": r.teacher_name or "",
        })
    return {"total": len(_query_records(db, start_date, end_date, room)), "records": out}


@router.get("/export")
def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    room: Optional[str] = None,
    user: CurrentUser = Depends(require_admin),
    db: Session = Depends(get_db),
):
    records = _query_records(db, start_date, end_date, room)
    students = {s.id: s for s in db.query(Student).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Room Inspections"

    headers = [
        "Date", "Member ID", "Student Name", "Class", "Section", "Room Number",
        "Bed Arrangement", "Cupboard", "Cleanliness", "Blanket Folded",
        "Remarks", "Recorded By",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in records:
        s = students.get(r.student_id)
        ws.append([
            r.date,
            s.member_id if s else "",
            s.name if s else "(deleted student)",
            s.class_ if s else "",
            s.section if s else "",
            s.room_number if s else "",
            "Yes" if r.bed_arrangement else "No",
            "Yes" if r.cupboard else "No",
            "Yes" if r.cleanliness else "No",
            "Yes" if r.blanket_folded else "No",
            r.remarks or "",
            r.teacher_name or "",
        ])

    widths = [12, 14, 22, 10, 10, 12, 16, 12, 12, 15, 28, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    db.add(AuditLog(
        user_id=user.id, actor_name=user.full_name, action="export_room_inspections",
        details=f"{len(records)} records ({start_date or 'all'} to {end_date or 'all'})",
    ))
    db.commit()

    stamp = datetime.utcnow().strftime("%Y%m%d")
    filename = f"room_inspections_{stamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
