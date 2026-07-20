"""Builds templates/student_import_template.xlsx - run with:
    python3 build_template.py
(openpyxl is required; it's already in backend/requirements.txt)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Students"

HEADERS = ["Member ID", "Name", "Password", "Phone Number"]
WIDTHS = [16, 28, 18, 18]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
BODY_FONT = Font(name="Arial", size=11)
EXAMPLE_FONT = Font(name="Arial", size=11, italic=True, color="808080")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

for col, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = WIDTHS[col - 1]
ws.row_dimensions[1].height = 22

# One example row showing the expected format - admins should replace/delete this row.
example = ["H001", "Aarav Sharma", "MyPass123", "9876543210"]
for col, val in enumerate(example, start=1):
    cell = ws.cell(row=2, column=col, value=val)
    cell.font = EXAMPLE_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")

# A few blank, bordered rows ready for real data entry.
for r in range(3, 23):
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=r, column=col)
        cell.font = BODY_FONT
        cell.border = BORDER

ws.freeze_panes = "A2"

# ---------------- Instructions sheet ----------------
info = wb.create_sheet("Instructions")
info.column_dimensions["A"].width = 100
lines = [
    ("Gurukul Rewards - Student Import Template", True),
    ("", False),
    ("How to use this file:", True),
    ("1. Fill in one row per student on the 'Students' sheet. Delete the example row (row 2) before importing, or just overwrite it.", False),
    ("2. Member ID - required. A unique ID for the student (e.g. their roll/admission code, like H001). This is also their login username.", False),
    ("3. Name - required. The student's full name.", False),
    ("4. Password - optional. If you set one, the student will log in with this exact password.", False),
    ("   If left blank, a default password is generated automatically: Gurukul@<MemberID> (e.g. Gurukul@H001).", False),
    ("5. Phone Number - optional. Digits only (no spaces, dashes, or country code symbols).", False),
    ("", False),
    ("Importing:", True),
    ("- Sign in as an Admin, go to the Admin dashboard -> Excel Import tab, and upload this file (.xlsx or save as .csv).", False),
    ("- Existing students (matched by Member ID) are updated; new Member IDs are added as new students.", False),
    ("- If you also want extra fields (Class, Section, House, Room Number, etc.), you can add more columns with those exact header names -", False),
    ("  the importer recognizes: Admission No, Class, Section, House, Room Number, Gender, Hostel, Photo URL, Date of Birth, Parent Name, Parent Phone, Email, Blood Group, Address, Joining Date.", False),
]
r = 1
for text, bold in lines:
    cell = info.cell(row=r, column=1, value=text)
    cell.font = Font(name="Arial", size=12 if bold else 11, bold=bold)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

wb.save("student_import_template.xlsx")
print("Wrote student_import_template.xlsx")
